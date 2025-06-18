# management/commands/enviar_reporte_mensual.py - Versión con WhatsApp integrado
import os
import calendar
from datetime import datetime, timedelta
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.db.models import Sum, Count, Q, Max
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

from registros.models import Registro, Equipo, Operador, ReporteGenerado, WhatsAppContact, WhatsAppMessage
from combustible.storage_backends import ReportesStorage, get_file_url
from whatsaap_service import WhatsAppReportService

logger = logging.getLogger(__name__)

class Command(BaseCommand):
    help = 'Envía reporte mensual por email y WhatsApp con integración completa'

    def add_arguments(self, parser):
        parser.add_argument('--mes', type=int, help='Mes específico (1-12)')
        parser.add_argument('--año', type=int, help='Año específico')
        parser.add_argument('--email', type=str, help='Email específico')
        parser.add_argument('--whatsapp', type=str, help='Número WhatsApp específico')
        parser.add_argument('--test', action='store_true', help='Modo test')
        parser.add_argument('--save-to-spaces', action='store_true', default=True, help='Guardar Excel en Spaces')
        parser.add_argument('--send-email', action='store_true', default=True, help='Enviar por email')
        parser.add_argument('--send-whatsapp', action='store_true', default=True, help='Enviar por WhatsApp')
        parser.add_argument('--whatsapp-only', action='store_true', help='Solo enviar por WhatsApp')

    def handle(self, *args, **options):
        try:
            # Determinar el período del reporte
            if options['mes'] and options['año']:
                año = options['año']
                mes = options['mes']
            else:
                # Mes anterior por defecto
                fecha_actual = datetime.now()
                if fecha_actual.month == 1:
                    mes = 12
                    año = fecha_actual.year - 1
                else:
                    mes = fecha_actual.month - 1
                    año = fecha_actual.year

            self.stdout.write(
                self.style.SUCCESS(f'📊 Generando reporte para {calendar.month_name[mes]} {año}...')
            )

            # Generar datos del reporte
            datos_reporte = self.generar_datos_reporte(año, mes)
            
            if options['test']:
                self.mostrar_estadisticas_consola(datos_reporte, año, mes)
                if options['send_whatsapp'] or options['whatsapp_only']:
                    self.test_whatsapp_functionality()
                return

            # Generar archivo Excel
            excel_buffer, excel_filename = self.generar_excel(datos_reporte, año, mes)
            
            # Guardar en Spaces si está habilitado
            reporte_obj = None
            excel_url = None
            if options['save_to_spaces']:
                reporte_obj = self.guardar_excel_en_spaces(
                    excel_buffer, excel_filename, datos_reporte, año, mes
                )
                if reporte_obj and reporte_obj.archivo_excel:
                    excel_url = reporte_obj.archivo_url

            # Determinar métodos de envío
            send_email = options['send_email'] and not options['whatsapp_only']
            send_whatsapp = options['send_whatsapp'] or options['whatsapp_only']

            # Enviar por correo
            if send_email:
                self.enviar_correo(
                    datos_reporte, excel_buffer, excel_filename, año, mes, 
                    options.get('email'), reporte_obj
                )

            # Enviar por WhatsApp
            if send_whatsapp:
                self.enviar_whatsapp(
                    datos_reporte, excel_url, excel_filename, año, mes,
                    options.get('whatsapp'), reporte_obj
                )
            
            # Marcar como enviado si se guardó en Spaces
            if reporte_obj:
                destinatarios_email = self.get_destinatarios_email(options.get('email')) if send_email else []
                destinatarios_whatsapp = self.get_destinatarios_whatsapp(options.get('whatsapp')) if send_whatsapp else []
                todos_destinatarios = destinatarios_email + [f"WhatsApp: {num}" for num in destinatarios_whatsapp]
                reporte_obj.marcar_como_enviado(todos_destinatarios)
            
            self.stdout.write(
                self.style.SUCCESS('✅ Reporte mensual enviado exitosamente!')
            )

        except Exception as e:
            logger.error(f"Error generando reporte: {str(e)}")
            self.stdout.write(
                self.style.ERROR(f'❌ Error al generar reporte: {str(e)}')
            )

    def generar_datos_reporte(self, año, mes):
        """Genera todos los datos necesarios para el reporte"""
        # [Código anterior de generación de datos - se mantiene igual]
        
        # Fechas del período
        primer_dia = datetime(año, mes, 1)
        ultimo_dia = datetime(año, mes, calendar.monthrange(año, mes)[1], 23, 59, 59)
        
        logger.info(f"Generando reporte para período: {primer_dia} - {ultimo_dia}")
        
        # Registros del mes con optimización de consultas
        registros_mes = Registro.objects.filter(
            fecha_hora__gte=primer_dia,
            fecha_hora__lte=ultimo_dia
        ).select_related('idEquipo', 'idOperador').prefetch_related().order_by('-fecha_hora')

        # Estadísticas generales
        total_registros = registros_mes.count()
        total_litros = registros_mes.aggregate(total=Sum('Litros'))['total'] or Decimal('0')
        total_gastado = sum([registro.total_costo for registro in registros_mes])

        # Estadísticas por equipo
        equipos_stats = {}
        for equipo in Equipo.objects.all().prefetch_related('registro_set'):
            registros_equipo = registros_mes.filter(idEquipo=equipo)
            
            if registros_equipo.exists():
                litros_equipo = registros_equipo.aggregate(total=Sum('Litros'))['total'] or Decimal('0')
                gasto_equipo = sum([reg.total_costo for reg in registros_equipo])
                ultimo_registro = registros_equipo.first()
                operadores_equipo = list(set([reg.idOperador.nombre for reg in registros_equipo]))
                
                equipos_stats[equipo.placa] = {
                    'equipo': equipo,
                    'total_litros': litros_equipo,
                    'total_gastado': gasto_equipo,
                    'num_registros': registros_equipo.count(),
                    'operadores': operadores_equipo,
                    'ultimo_registro': ultimo_registro,
                    'promedio_litros': litros_equipo / registros_equipo.count() if registros_equipo.count() > 0 else 0,
                }

        # Operadores que NO cargaron combustible
        operadores_activos = set(registros_mes.values_list('idOperador', flat=True))
        operadores_inactivos = Operador.objects.exclude(id__in=operadores_activos)

        # Estadísticas por operador
        operadores_stats = {}
        for operador in Operador.objects.all():
            registros_operador = registros_mes.filter(idOperador=operador)
            
            if registros_operador.exists():
                litros_operador = registros_operador.aggregate(total=Sum('Litros'))['total'] or Decimal('0')
                gasto_operador = sum([reg.total_costo for reg in registros_operador])
                equipos_usados = list(set([reg.idEquipo.placa for reg in registros_operador]))
                
                operadores_stats[operador.nombre] = {
                    'operador': operador,
                    'total_litros': litros_operador,
                    'total_gastado': gasto_operador,
                    'num_registros': registros_operador.count(),
                    'equipos_usados': equipos_usados,
                    'promedio_litros': litros_operador / registros_operador.count() if registros_operador.count() > 0 else 0,
                }

        # Top rankings
        top_equipos = sorted(equipos_stats.items(), key=lambda x: x[1]['total_litros'], reverse=True)[:5]
        top_operadores = sorted(operadores_stats.items(), key=lambda x: x[1]['num_registros'], reverse=True)[:5]

        # Estadísticas adicionales
        registros_con_foto = registros_mes.exclude(photo_tiket__isnull=True).exclude(photo_tiket='').count()
        porcentaje_con_foto = (registros_con_foto / total_registros * 100) if total_registros > 0 else 0

        return {
            'año': año,
            'mes': mes,
            'nombre_mes': calendar.month_name[mes],
            'primer_dia': primer_dia,
            'ultimo_dia': ultimo_dia,
            'registros': list(registros_mes),
            'total_registros': total_registros,
            'total_litros': total_litros,
            'total_gastado': total_gastado,
            'equipos_stats': equipos_stats,
            'operadores_stats': operadores_stats,
            'operadores_inactivos': list(operadores_inactivos),
            'top_equipos': top_equipos,
            'top_operadores': top_operadores,
            'promedio_diario': total_gastado / calendar.monthrange(año, mes)[1] if total_gastado > 0 else 0,
            'promedio_litros_registro': total_litros / total_registros if total_registros > 0 else 0,
            'registros_con_foto': registros_con_foto,
            'porcentaje_con_foto': porcentaje_con_foto,
        }

    def generar_excel(self, datos, año, mes):
        """Genera el archivo Excel - código anterior se mantiene"""
        # [Código anterior de generación de Excel]
        wb = Workbook()
        
        # Generar nombre de archivo único
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reporte_combustible_{año}_{mes:02d}_{timestamp}.xlsx"
        
        # [Resto del código de generación de Excel...]
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer, filename

    def guardar_excel_en_spaces(self, excel_buffer, filename, datos_reporte, año, mes):
        """Guarda el archivo Excel en DigitalOcean Spaces - código anterior se mantiene"""
        try:
            # Crear objeto ReporteGenerado
            reporte = ReporteGenerado.objects.create(
                nombre=f"Reporte Mensual - {datos_reporte['nombre_mes']} {año}",
                tipo='mensual',
                fecha_inicio=datos_reporte['primer_dia'].date(),
                fecha_fin=datos_reporte['ultimo_dia'].date(),
                total_registros=datos_reporte['total_registros'],
                total_litros=datos_reporte['total_litros'],
                total_gastado=datos_reporte['total_gastado']
            )
            
            # Guardar archivo en Spaces
            excel_content = ContentFile(excel_buffer.getvalue(), name=filename)
            reporte.archivo_excel.save(filename, excel_content)
            
            logger.info(f"📁 Excel guardado en Spaces: {reporte.archivo_url}")
            return reporte
            
        except Exception as e:
            logger.error(f"Error guardando Excel en Spaces: {str(e)}")
            return None

    def enviar_whatsapp(self, datos, excel_url, filename, año, mes, numero_especifico=None, reporte_obj=None):
        """Envía el reporte por WhatsApp"""
        self.stdout.write("📱 Enviando reporte por WhatsApp...")
        
        try:
            whatsapp_service = WhatsAppReportService()
            destinatarios = self.get_destinatarios_whatsapp(numero_especifico)
            
            if not destinatarios:
                self.stdout.write(
                    self.style.WARNING("⚠️ No hay destinatarios de WhatsApp configurados")
                )
                return
            
            total_enviados = 0
            total_errores = 0
            
            for numero in destinatarios:
                try:
                    # Obtener contacto para personalización
                    contact = None
                    try:
                        contact = WhatsAppContact.objects.get(phone_number__contains=numero[-10:])
                    except WhatsAppContact.DoesNotExist:
                        pass
                    
                    # Enviar resumen y archivo
                    if excel_url:
                        results = whatsapp_service.send_report_with_excel(
                            numero, datos, excel_url, filename
                        )
                    else:
                        # Solo enviar resumen si no hay URL del Excel
                        results = [whatsapp_service.send_monthly_report_summary(numero, datos)]
                    
                    # Verificar si se enviaron correctamente
                    success_count = sum(1 for result in results if result.get('success', False))
                    
                    if success_count > 0:
                        total_enviados += 1
                        
                        # Registrar mensajes enviados
                        self._registrar_mensajes_whatsapp(contact, results, reporte_obj)
                        
                        self.stdout.write(f"   ✅ Enviado a {numero} ({success_count}/{len(results)} mensajes)")
                        
                        # Enviar alertas específicas si aplica
                        if datos.get('operadores_inactivos'):
                            whatsapp_service.send_alert_inactive_operators(
                                numero, datos['operadores_inactivos'], datos['nombre_mes']
                            )
                        
                    else:
                        total_errores += 1
                        self.stdout.write(f"   ❌ Error enviando a {numero}")
                
                except Exception as e:
                    total_errores += 1
                    logger.error(f"Error enviando WhatsApp a {numero}: {e}")
                    self.stdout.write(f"   ❌ Error enviando a {numero}: {str(e)}")
            
            self.stdout.write(
                self.style.SUCCESS(f"📱 WhatsApp: {total_enviados} enviados, {total_errores} errores")
            )
            
        except Exception as e:
            logger.error(f"Error general enviando WhatsApp: {e}")
            self.stdout.write(
                self.style.ERROR(f"❌ Error enviando WhatsApp: {str(e)}")
            )

    def _registrar_mensajes_whatsapp(self, contact, results, reporte_obj):
        """Registra los mensajes enviados en la base de datos"""
        try:
            for result in results:
                if result.get('success') and result.get('data'):
                    message_data = result['data']
                    messages = message_data.get('messages', [])
                    
                    for msg in messages:
                        WhatsAppMessage.objects.create(
                            contact=contact,
                            message_type='text',  # Simplificado por ahora
                            content='Reporte mensual enviado',
                            whatsapp_message_id=msg.get('id'),
                            status='sent',
                            reporte=reporte_obj
                        )
                        
                        # Actualizar último mensaje enviado del contacto
                        if contact:
                            contact.last_message_sent = timezone.now()
                            contact.save(update_fields=['last_message_sent'])
        
        except Exception as e:
            logger.error(f"Error registrando mensajes WhatsApp: {e}")

    def get_destinatarios_whatsapp(self, numero_especifico=None):
        """Obtiene la lista de destinatarios de WhatsApp"""
        if numero_especifico:
            return [numero_especifico]
        
        # Obtener de contactos activos que reciben reportes
        contactos = WhatsAppContact.objects.filter(
            active=True,
            receive_monthly_reports=True
        ).values_list('phone_number', flat=True)
        
        return list(contactos)

    def get_destinatarios_email(self, email_especifico=None):
        """Obtiene la lista de destinatarios de email"""
        if email_especifico:
            return [email_especifico]
        
        return getattr(settings, 'REPORTES_EMAIL_DESTINATARIOS', [
            'admin@empresa.com',
            'gerencia@empresa.com'
        ])

    def enviar_correo(self, datos, excel_buffer, filename, año, mes, email_especifico=None, reporte_obj=None):
        """Envía el correo con el reporte - código anterior se mantiene"""
        destinatarios = self.get_destinatarios_email(email_especifico)
        
        # Contexto del email
        contexto_email = {
            'datos': datos,
            'año': año,
            'mes': mes,
            'reporte_url': reporte_obj.archivo_url if reporte_obj else None,
        }
        
        # Generar HTML del email
        html_content = render_to_string('emails/reporte_mensual.html', contexto_email)
        
        # Crear email
        asunto = f"📊 Reporte Mensual de Combustible - {datos['nombre_mes']} {año}"
        
        email = EmailMessage(
            subject=asunto,
            body=html_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=destinatarios,
        )
        email.content_subtype = "html"
        
        # Adjuntar Excel
        email.attach(filename, excel_buffer.getvalue(), 
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Enviar
        email.send()
        
        logger.info(f"📧 Email enviado a: {', '.join(destinatarios)}")
        self.stdout.write(
            self.style.SUCCESS(f'📧 Email enviado a: {", ".join(destinatarios)}')
        )

    def test_whatsapp_functionality(self):
        """Prueba la funcionalidad de WhatsApp en modo test"""
        self.stdout.write("\n🧪 PROBANDO FUNCIONALIDAD DE WHATSAPP")
        self.stdout.write("-" * 50)
        
        try:
            from .whatsapp_service import WhatsAppBusinessService
            service = WhatsAppBusinessService()
            
            # Verificar configuración
            if not service.access_token or not service.phone_number_id:
                self.stdout.write(
                    self.style.ERROR("❌ WhatsApp no está configurado correctamente")
                )
                return
            
            self.stdout.write("✅ Configuración de WhatsApp válida")
            
            # Contar contactos activos
            contactos_activos = WhatsAppContact.objects.filter(active=True).count()
            contactos_reportes = WhatsAppContact.objects.filter(
                active=True, receive_monthly_reports=True
            ).count()
            
            self.stdout.write(f"📱 Contactos activos: {contactos_activos}")
            self.stdout.write(f"📊 Contactos que reciben reportes: {contactos_reportes}")
            
            if contactos_reportes == 0:
                self.stdout.write(
                    self.style.WARNING("⚠️ No hay contactos configurados para recibir reportes")
                )
            
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f"❌ Error probando WhatsApp: {e}")
            )

    def mostrar_estadisticas_consola(self, datos, año, mes):
        """Muestra estadísticas en consola con información de WhatsApp"""
        
        print(f"\n{'='*70}")
        print(f"📊 REPORTE MENSUAL DE COMBUSTIBLE - {datos['nombre_mes'].upper()} {año}")
        print(f"{'='*70}")
        
        print(f"\n📈 ESTADÍSTICAS GENERALES:")
        print(f"   • Total de registros: {datos['total_registros']}")
        print(f"   • Registros con foto: {datos['registros_con_foto']} ({datos['porcentaje_con_foto']:.1f}%)")
        print(f"   • Total de litros: {datos['total_litros']:.2f}L")
        print(f"   • Total gastado: ${datos['total_gastado']:.2f}")
        print(f"   • Promedio diario: ${datos['promedio_diario']:.2f}")
        
        print(f"\n🚛 TOP 5 EQUIPOS (por consumo):")
        for i, (placa, stats) in enumerate(datos['top_equipos'], 1):
            print(f"   {i}. {placa} - {stats['total_litros']:.2f}L (${stats['total_gastado']:.2f})")
        
        print(f"\n👥 TOP 5 OPERADORES (por actividad):")
        for i, (nombre, stats) in enumerate(datos['top_operadores'], 1):
            print(f"   {i}. {nombre} - {stats['num_registros']} registros ({stats['total_litros']:.2f}L)")
        
        if datos['operadores_inactivos']:
            print(f"\n⚠️  OPERADORES SIN ACTIVIDAD:")
            for operador in datos['operadores_inactivos']:
                print(f"   • {operador.nombre} ({operador.email})")
        
        print(f"\n💾 INTEGRACIÓN CON SPACES:")
        print(f"   • Fotos de tickets almacenadas: ✅")
        print(f"   • Reportes Excel en la nube: ✅")
        print(f"   • URLs públicas disponibles: ✅")
        
        # Estadísticas de WhatsApp
        try:
            contactos_whatsapp = WhatsAppContact.objects.filter(active=True).count()
            contactos_reportes = WhatsAppContact.objects.filter(
                active=True, receive_monthly_reports=True
            ).count()
            
            print(f"\n📱 INTEGRACIÓN CON WHATSAPP:")
            print(f"   • Contactos activos: {contactos_whatsapp}")
            print(f"   • Contactos que reciben reportes: {contactos_reportes}")
            print(f"   • Envío automático de resúmenes: ✅")
            print(f"   • Archivos Excel por WhatsApp: ✅")
            print(f"   • Alertas de operadores inactivos: ✅")
            
        except Exception as e:
            print(f"\n📱 WHATSAPP: Error obteniendo estadísticas - {e}")
        
        print(f"\n{'='*70}")

# === COMANDO ADICIONAL PARA GESTIÓN DE WHATSAPP ===

# management/commands/manage_whatsapp_contacts.py
from django.core.management.base import BaseCommand
from registros.models import WhatsAppContact, Operador
from whatsaap_service import WhatsAppBusinessService

class Command(BaseCommand):
    help = 'Gestiona contactos de WhatsApp para reportes'

    def add_arguments(self, parser):
        parser.add_argument('--add', type=str, help='Agregar contacto: nombre,numero,rol')
        parser.add_argument('--list', action='store_true', help='Listar contactos')
        parser.add_argument('--test', type=str, help='Enviar mensaje de prueba a número')
        parser.add_argument('--sync', action='store_true', help='Sincronizar con operadores')
        parser.add_argument('--enable-reports', type=str, help='Habilitar reportes para número')
        parser.add_argument('--disable-reports', type=str, help='Deshabilitar reportes para número')

    def handle(self, *args, **options):
        if options['add']:
            self.add_contact(options['add'])
        elif options['list']:
            self.list_contacts()
        elif options['test']:
            self.test_message(options['test'])
        elif options['sync']:
            self.sync_with_operators()
        elif options['enable_reports']:
            self.toggle_reports(options['enable_reports'], True)
        elif options['disable_reports']:
            self.toggle_reports(options['disable_reports'], False)
        else:
            self.stdout.write("❌ Especifica una acción: --add, --list, --test, --sync, --enable-reports, --disable-reports")

    def add_contact(self, contact_data):
        """Agrega un nuevo contacto"""
        try:
            parts = contact_data.split(',')
            if len(parts) != 3:
                self.stdout.write(self.style.ERROR("❌ Formato: nombre,numero,rol"))
                return
            
            name, number, role = [part.strip() for part in parts]
            
            contact, created = WhatsAppContact.objects.get_or_create(
                phone_number=number,
                defaults={
                    'name': name,
                    'role': role,
                    'receive_monthly_reports': True,
                }
            )
            
            if created:
                self.stdout.write(self.style.SUCCESS(f"✅ Contacto agregado: {name}"))
            else:
                self.stdout.write(self.style.WARNING(f"⚠️ Contacto ya existe: {name}"))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error agregando contacto: {e}"))

    def list_contacts(self):
        """Lista todos los contactos"""
        contacts = WhatsAppContact.objects.all().order_by('role', 'name')
        
        if not contacts:
            self.stdout.write("📱 No hay contactos registrados")
            return
        
        self.stdout.write("\n📱 CONTACTOS DE WHATSAPP:")
        self.stdout.write("-" * 60)
        
        for contact in contacts:
            status = "✅" if contact.active else "❌"
            reports = "📊" if contact.receive_monthly_reports else "🚫"
            
            self.stdout.write(
                f"{status} {contact.name} ({contact.get_role_display()}) - {contact.phone_number} {reports}"
            )

    def test_message(self, number):
        """Envía mensaje de prueba"""
        try:
            service = WhatsAppBusinessService()
            
            message = "🤖 *Mensaje de Prueba*\n\nSistema de Combustible funcionando correctamente ✅\n\nEste es un mensaje de prueba del sistema de reportes automáticos."
            
            result = service.send_text_message(number, message)
            
            if result['success']:
                self.stdout.write(self.style.SUCCESS(f"✅ Mensaje de prueba enviado a {number}"))
            else:
                self.stdout.write(self.style.ERROR(f"❌ Error: {result.get('error', 'Error desconocido')}"))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error enviando prueba: {e}"))

    def sync_with_operators(self):
        """Sincroniza contactos con operadores del sistema"""
        operadores = Operador.objects.exclude(movil__isnull=True).exclude(movil='')
        
        synced = 0
        for operador in operadores:
            contact, created = WhatsAppContact.objects.get_or_create(
                phone_number=operador.movil,
                defaults={
                    'name': operador.nombre,
                    'role': 'operator',
                    'operador': operador,
                    'receive_monthly_reports': False,  # Operadores no reciben reportes por defecto
                    'receive_alerts': True,
                }
            )
            
            if created:
                synced += 1
        
        self.stdout.write(self.style.SUCCESS(f"🔄 Sincronización completada: {synced} contactos nuevos"))

    def toggle_reports(self, number, enable):
        """Habilita/deshabilita reportes para un contacto"""
        try:
            contact = WhatsAppContact.objects.get(phone_number__contains=number[-10:])
            contact.receive_monthly_reports = enable
            contact.save()
            
            action = "habilitados" if enable else "deshabilitados"
            self.stdout.write(self.style.SUCCESS(f"✅ Reportes {action} para {contact.name}"))
            
        except WhatsAppContact.DoesNotExist:
            self.stdout.write(self.style.ERROR(f"❌ Contacto no encontrado: {number}"))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"❌ Error: {e}"))

# === PLANTILLAS DE WHATSAPP ===

WHATSAPP_TEMPLATES = {
    'monthly_report_notification': {
        'name': 'monthly_report_notification',
        'language': 'es_MX',
        'components': [
            {
                'type': 'HEADER',
                'format': 'TEXT',
                'text': '📊 Reporte Mensual Disponible'
            },
            {
                'type': 'BODY',
                'text': 'Hola {{1}}, el reporte mensual de combustible de {{2}} {{3}} está listo.\n\n• Total gastado: ${{4}}\n• Total litros: {{5}}L\n• Registros: {{6}}\n\nEl archivo Excel se envía a continuación.'
            },
            {
                'type': 'FOOTER',
                'text': 'Sistema de Gestión de Combustible'
            }
        ]
    },
    'inactive_operators_alert': {
        'name': 'inactive_operators_alert',
        'language': 'es_MX',
        'components': [
            {
                'type': 'HEADER',
                'format': 'TEXT',
                'text': '⚠️ Alerta: Operadores Inactivos'
            },
            {
                'type': 'BODY',
                'text': 'Se detectaron {{1}} operadores sin actividad en {{2}}.\n\nSe recomienda contactarlos para verificar su estado.'
            }
        ]
    }
}