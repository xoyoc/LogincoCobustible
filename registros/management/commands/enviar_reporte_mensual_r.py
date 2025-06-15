# management/commands/enviar_reporte_mensual.py
import os
import calendar
from datetime import datetime, timedelta
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.db.models import Sum, Count, Q, Max
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from registros.models import Registro, Equipo, Operador  # Ajusta la importación según tu estructura


class Command(BaseCommand):
    help = 'Envía reporte mensual de combustible por correo electrónico'

    def add_arguments(self, parser):
        parser.add_argument(
            '--mes',
            type=int,
            help='Mes específico (1-12). Si no se especifica, usa el mes anterior'
        )
        parser.add_argument(
            '--año',
            type=int,
            help='Año específico. Si no se especifica, usa el año actual'
        )
        parser.add_argument(
            '--email',
            type=str,
            help='Email específico para enviar el reporte. Si no se especifica, usa los configurados'
        )
        parser.add_argument(
            '--test',
            action='store_true',
            help='Modo de prueba - no envía email, solo muestra estadísticas'
        )

    def handle(self, *args, **options):
        try:
            # Determinar el período del reporte
            if options['mes'] and options['año']:
                año = options['año']
                mes = options['mes']
            else:
                # Mes anterior por defecto
                fecha_actual = datetime.now()
                if fecha_actual.month == 1:
                    mes = 12
                    año = fecha_actual.year - 1
                else:
                    mes = fecha_actual.month - 1
                    año = fecha_actual.year

            self.stdout.write(
                self.style.SUCCESS(f'Generando reporte para {calendar.month_name[mes]} {año}...')
            )

            # Generar datos del reporte
            datos_reporte = self.generar_datos_reporte(año, mes)
            
            if options['test']:
                self.mostrar_estadisticas_consola(datos_reporte, año, mes)
                return

            # Generar archivo Excel
            archivo_excel = self.generar_excel(datos_reporte, año, mes)
            
            # Enviar por correo
            self.enviar_correo(datos_reporte, archivo_excel, año, mes, options.get('email'))
            
            self.stdout.write(
                self.style.SUCCESS(f'Reporte mensual enviado exitosamente!')
            )

        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Error al generar reporte: {str(e)}')
            )

    def generar_datos_reporte(self, año, mes):
        """Genera todos los datos necesarios para el reporte"""
        
        # Fechas del período
        primer_dia = datetime(año, mes, 1)
        ultimo_dia = datetime(año, mes, calendar.monthrange(año, mes)[1], 23, 59, 59)
        
        # Registros del mes
        registros_mes = Registro.objects.filter(
            fecha_hora__gte=primer_dia,
            fecha_hora__lte=ultimo_dia
        ).select_related('idEquipo', 'idOperador').order_by('-fecha_hora')

        # Estadísticas generales
        total_registros = registros_mes.count()
        total_litros = registros_mes.aggregate(
            total=Sum('Litros')
        )['total'] or Decimal('0')
        
        total_gastado = sum([
            registro.Litros * registro.costolitro 
            for registro in registros_mes
        ])

        # Estadísticas por equipo
        equipos_stats = {}
        for equipo in Equipo.objects.all():
            registros_equipo = registros_mes.filter(idEquipo=equipo)
            
            if registros_equipo.exists():
                litros_equipo = registros_equipo.aggregate(
                    total=Sum('Litros')
                )['total'] or Decimal('0')
                
                gasto_equipo = sum([
                    reg.Litros * reg.costolitro 
                    for reg in registros_equipo
                ])
                
                ultimo_registro = registros_equipo.first()
                operadores_equipo = list(set([
                    reg.idOperador.nombre 
                    for reg in registros_equipo
                ]))
                
                equipos_stats[equipo.placa] = {
                    'equipo': equipo,
                    'total_litros': litros_equipo,
                    'total_gastado': gasto_equipo,
                    'num_registros': registros_equipo.count(),
                    'operadores': operadores_equipo,
                    'ultimo_registro': ultimo_registro,
                    'promedio_litros': litros_equipo / registros_equipo.count() if registros_equipo.count() > 0 else 0
                }

        # Operadores que NO cargaron combustible
        operadores_activos = set(registros_mes.values_list('idOperador', flat=True))
        operadores_inactivos = Operador.objects.exclude(
            id__in=operadores_activos
        )

        # Estadísticas por operador
        operadores_stats = {}
        for operador in Operador.objects.all():
            registros_operador = registros_mes.filter(idOperador=operador)
            
            if registros_operador.exists():
                litros_operador = registros_operador.aggregate(
                    total=Sum('Litros')
                )['total'] or Decimal('0')
                
                gasto_operador = sum([
                    reg.Litros * reg.costolitro 
                    for reg in registros_operador
                ])
                
                equipos_usados = list(set([
                    reg.idEquipo.placa 
                    for reg in registros_operador
                ]))
                
                operadores_stats[operador.nombre] = {
                    'operador': operador,
                    'total_litros': litros_operador,
                    'total_gastado': gasto_operador,
                    'num_registros': registros_operador.count(),
                    'equipos_usados': equipos_usados,
                    'promedio_litros': litros_operador / registros_operador.count() if registros_operador.count() > 0 else 0
                }

        # Top 5 equipos con más consumo
        top_equipos = sorted(
            equipos_stats.items(), 
            key=lambda x: x[1]['total_litros'], 
            reverse=True
        )[:5]

        # Top 5 operadores más activos
        top_operadores = sorted(
            operadores_stats.items(), 
            key=lambda x: x[1]['num_registros'], 
            reverse=True
        )[:5]

        # Estadísticas por día del mes
        stats_diarias = {}
        for dia in range(1, calendar.monthrange(año, mes)[1] + 1):
            fecha_dia = datetime(año, mes, dia)
            registros_dia = registros_mes.filter(
                fecha_hora__date=fecha_dia.date()
            )
            
            if registros_dia.exists():
                litros_dia = registros_dia.aggregate(
                    total=Sum('Litros')
                )['total'] or Decimal('0')
                
                gasto_dia = sum([
                    reg.Litros * reg.costolitro 
                    for reg in registros_dia
                ])
                
                stats_diarias[dia] = {
                    'fecha': fecha_dia,
                    'registros': registros_dia.count(),
                    'litros': litros_dia,
                    'gasto': gasto_dia
                }

        return {
            'año': año,
            'mes': mes,
            'nombre_mes': calendar.month_name[mes],
            'primer_dia': primer_dia,
            'ultimo_dia': ultimo_dia,
            'registros': list(registros_mes),
            'total_registros': total_registros,
            'total_litros': total_litros,
            'total_gastado': total_gastado,
            'equipos_stats': equipos_stats,
            'operadores_stats': operadores_stats,
            'operadores_inactivos': list(operadores_inactivos),
            'top_equipos': top_equipos,
            'top_operadores': top_operadores,
            'stats_diarias': stats_diarias,
            'promedio_diario': total_gastado / calendar.monthrange(año, mes)[1] if total_gastado > 0 else 0,
            'promedio_litros_registro': total_litros / total_registros if total_registros > 0 else 0
        }

    def generar_excel(self, datos, año, mes):
        """Genera el archivo Excel con todos los registros y estadísticas"""
        
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === HOJA 1: RESUMEN EJECUTIVO ===
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"
        
        # Título
        ws_resumen.merge_cells('A1:F1')
        ws_resumen['A1'] = f"REPORTE MENSUAL DE COMBUSTIBLE - {datos['nombre_mes'].upper()} {año}"
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        
        # Información general
        row = 3
        info_general = [
            ["Período:", f"{datos['primer_dia'].strftime('%d/%m/%Y')} - {datos['ultimo_dia'].strftime('%d/%m/%Y')}"],
            ["Total de Registros:", datos['total_registros']],
            ["Total de Litros:", f"{datos['total_litros']:.2f} L"],
            ["Total Gastado:", f"${datos['total_gastado']:.2f}"],
            ["Promedio Diario:", f"${datos['promedio_diario']:.2f}"],
            ["Promedio por Registro:", f"{datos['promedio_litros_registro']:.2f} L"]
        ]
        
        for info in info_general:
            ws_resumen[f'A{row}'] = info[0]
            ws_resumen[f'A{row}'].font = Font(bold=True)
            ws_resumen[f'B{row}'] = info[1]
            row += 1
        
        # Top 5 Equipos
        row += 2
        ws_resumen[f'A{row}'] = "TOP 5 EQUIPOS (Consumo de Litros)"
        ws_resumen[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Placa", "Marca/Modelo", "Litros", "Gasto", "Registros"]
        for col, header in enumerate(headers, 1):
            cell = ws_resumen.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for placa, stats in datos['top_equipos']:
            ws_resumen[f'A{row}'] = placa
            ws_resumen[f'B{row}'] = f"{stats['equipo'].marca} {stats['equipo'].modelo}"
            ws_resumen[f'C{row}'] = f"{stats['total_litros']:.2f}"
            ws_resumen[f'D{row}'] = f"${stats['total_gastado']:.2f}"
            ws_resumen[f'E{row}'] = stats['num_registros']
            
            for col in range(1, 6):
                ws_resumen.cell(row=row, column=col).border = border
            row += 1
        
        # Operadores Inactivos
        if datos['operadores_inactivos']:
            row += 2
            ws_resumen[f'A{row}'] = "OPERADORES SIN ACTIVIDAD EN EL MES"
            ws_resumen[f'A{row}'].font = Font(bold=True, size=12, color="FF0000")
            row += 1
            
            for operador in datos['operadores_inactivos']:
                ws_resumen[f'A{row}'] = f"• {operador.nombre} ({operador.email})"
                row += 1
        
        # === HOJA 2: DETALLE DE REGISTROS ===
        ws_detalle = wb.create_sheet("Detalle de Registros")
        
        # Headers
        headers_detalle = [
            "Fecha/Hora", "Ticket", "Placa", "Marca", "Modelo", 
            "Operador", "Litros", "Costo/L", "Total", "Kilometraje"
        ]
        
        for col, header in enumerate(headers_detalle, 1):
            cell = ws_detalle.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Datos
        for row, registro in enumerate(datos['registros'], 2):
            ws_detalle[f'A{row}'] = registro.fecha_hora.strftime('%d/%m/%Y %H:%M')
            ws_detalle[f'B{row}'] = registro.numero_tiket
            ws_detalle[f'C{row}'] = registro.idEquipo.placa
            ws_detalle[f'D{row}'] = registro.idEquipo.marca
            ws_detalle[f'E{row}'] = registro.idEquipo.modelo
            ws_detalle[f'F{row}'] = registro.idOperador.nombre
            ws_detalle[f'G{row}'] = float(registro.Litros)
            ws_detalle[f'H{row}'] = float(registro.costolitro)
            ws_detalle[f'I{row}'] = float(registro.Litros * registro.costolitro)
            ws_detalle[f'J{row}'] = registro.kilometraje
            
            for col in range(1, 11):
                ws_detalle.cell(row=row, column=col).border = border
        
        # === HOJA 3: ESTADÍSTICAS POR EQUIPO ===
        ws_equipos = wb.create_sheet("Estadísticas por Equipo")
        
        headers_equipos = [
            "Placa", "Marca/Modelo", "Año", "Capacidad", 
            "Litros", "Gasto", "Registros", "Promedio/Registro", "Operadores"
        ]
        
        for col, header in enumerate(headers_equipos, 1):
            cell = ws_equipos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 2
        for placa, stats in datos['equipos_stats'].items():
            ws_equipos[f'A{row}'] = placa
            ws_equipos[f'B{row}'] = f"{stats['equipo'].marca} {stats['equipo'].modelo}"
            ws_equipos[f'C{row}'] = stats['equipo'].year
            ws_equipos[f'D{row}'] = f"{stats['equipo'].capacidad_tanque}L"
            ws_equipos[f'E{row}'] = f"{stats['total_litros']:.2f}"
            ws_equipos[f'F{row}'] = f"${stats['total_gastado']:.2f}"
            ws_equipos[f'G{row}'] = stats['num_registros']
            ws_equipos[f'H{row}'] = f"{stats['promedio_litros']:.2f}L"
            ws_equipos[f'I{row}'] = ", ".join(stats['operadores'])
            
            for col in range(1, 10):
                ws_equipos.cell(row=row, column=col).border = border
            row += 1
        
        # Ajustar ancho de columnas
        for ws in [ws_resumen, ws_detalle, ws_equipos]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    def enviar_correo(self, datos, archivo_excel, año, mes, email_especifico=None):
        """Envía el correo con el reporte"""
        
        # Emails destinatarios
        if email_especifico:
            destinatarios = [email_especifico]
        else:
            # Configurar en settings.py: REPORTES_EMAIL_DESTINATARIOS
            destinatarios = getattr(settings, 'REPORTES_EMAIL_DESTINATARIOS', [
                'zuly.becerra@loginco.com.mx',
                'xoyoc_l2@hotmail.com',
            ])
        
        # Generar HTML del email
        contexto_email = {
            'datos': datos,
            'año': año,
            'mes': mes
        }
        
        html_content = render_to_string('emails/reporte_mensual.html', contexto_email)
        
        # Crear email
        asunto = f"Reporte Mensual de Combustible - {datos['nombre_mes']} {año}"
        
        email = EmailMessage(
            subject=asunto,
            body=html_content,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=destinatarios,
        )
        email.content_subtype = "html"
        
        # Adjuntar Excel
        nombre_archivo = f"reporte_combustible_{año}_{mes:02d}.xlsx"
        email.attach(nombre_archivo, archivo_excel.getvalue(), 
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Enviar
        email.send()
        
        self.stdout.write(
            self.style.SUCCESS(f'Email enviado a: {", ".join(destinatarios)}')
        )

    def mostrar_estadisticas_consola(self, datos, año, mes):
        """Muestra las estadísticas en la consola para modo test"""
        
        print(f"\n{'='*60}")
        print(f"REPORTE MENSUAL DE COMBUSTIBLE - {datos['nombre_mes'].upper()} {año}")
        print(f"{'='*60}")
        
        print(f"\n📊 ESTADÍSTICAS GENERALES:")
        print(f"   • Total de registros: {datos['total_registros']}")
        print(f"   • Total de litros: {datos['total_litros']:.2f}L")
        print(f"   • Total gastado: ${datos['total_gastado']:.2f}")
        print(f"   • Promedio diario: ${datos['promedio_diario']:.2f}")
        
        print(f"\n🚛 TOP 5 EQUIPOS (por consumo):")
        for i, (placa, stats) in enumerate(datos['top_equipos'], 1):
            print(f"   {i}. {placa} - {stats['total_litros']:.2f}L (${stats['total_gastado']:.2f})")
        
        print(f"\n👥 TOP 5 OPERADORES (por actividad):")
        for i, (nombre, stats) in enumerate(datos['top_operadores'], 1):
            print(f"   {i}. {nombre} - {stats['num_registros']} registros ({stats['total_litros']:.2f}L)")
        
        if datos['operadores_inactivos']:
            print(f"\n⚠️  OPERADORES SIN ACTIVIDAD:")
            for operador in datos['operadores_inactivos']:
                print(f"   • {operador.nombre} ({operador.email})")
        
        print(f"\n{'='*60}")